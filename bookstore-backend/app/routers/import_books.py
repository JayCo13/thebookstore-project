"""
Bulk book import from XLSX files.
Parses the spreadsheet, auto-creates authors/categories, and creates books.
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
import re
import io
import logging
import unicodedata

from app.database import get_db, get_redis
from app.models.models import Book, Author, Category, book_authors, book_categories
from app.middleware.auth_middleware import require_admin
from app.models.models import User
from app.cache.redis_cache import RedisCache
from redis import Redis

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/books", tags=["Books Import"])


def _slugify(text: str) -> str:
    """Generate a URL-friendly slug from text."""
    if not text:
        return ""
    n = unicodedata.normalize('NFKD', text)
    ascii_text = ''.join(c for c in n if not unicodedata.combining(c))
    ascii_text = ascii_text.lower()
    allowed = []
    for ch in ascii_text:
        if ch.isalnum():
            allowed.append(ch)
        elif ch in [' ', '-', '_']:
            allowed.append('-')
    slug = ''.join(allowed)
    while '--' in slug:
        slug = slug.replace('--', '-')
    return slug.strip('-')


def _parse_features(features_str: str) -> dict:
    """
    Parse the Features column string to extract structured data.
    
    Example input:
    (Đặc điểm chung) Loại sản phẩm: S[Bìa mềm]; (Đặc điểm chung) Kích thước: T[15.5 x 23.5];
    (Sách) Tác giả: M[Mauro F. Guillén]; (Sách) Nhà Xuất Bản: M[NXB Thế Giới];
    (Sách) Số trang: O[360]; ...
    """
    result = {
        'author': None,
        'publisher': None,
        'pages': None,
        'height': None,
        'width': None,
        'translator': None
    }
    
    if not features_str:
        return result
    
    # Extract author: Tác giả: M[...] or Tác giả: E[...]
    author_match = re.search(r'Tác giả:\s*\w\[([^\]]+)\]', features_str)
    if author_match:
        result['author'] = author_match.group(1).strip()
    
    # Extract translator: Dịch giả: E[...]
    translator_match = re.search(r'Dịch giả:\s*\w\[([^\]]+)\]', features_str)
    if translator_match:
        result['translator'] = translator_match.group(1).strip()
    
    # Extract publisher: Nhà Xuất Bản: M[...]
    publisher_match = re.search(r'Nhà Xuất Bản:\s*\w\[([^\]]+)\]', features_str)
    if publisher_match:
        result['publisher'] = publisher_match.group(1).strip()
    
    # Extract pages: Số trang: O[...]
    pages_match = re.search(r'Số trang:\s*\w\[([^\]]+)\]', features_str)
    if pages_match:
        try:
            result['pages'] = int(pages_match.group(1).strip())
        except ValueError:
            pass
    
    # Extract dimensions: Kích thước: T[15.5 x 23.5]
    dim_match = re.search(r'Kích thước:\s*\w\[([^\]]+)\]', features_str)
    if dim_match:
        dims = dim_match.group(1).strip()
        parts = re.split(r'\s*x\s*', dims)
        if len(parts) >= 2:
            try:
                result['width'] = float(parts[0].strip())
                result['height'] = float(parts[1].strip())
            except ValueError:
                pass
    
    return result


def _parse_category(category_str: str) -> str:
    """
    Extract the most specific category from the path.
    e.g. "Sách Tiếng Việt///Theo Thể Loại///Kinh Tế/Kinh Doanh///Phân tích - Môi trường kinh tế"
    → "Phân tích - Môi trường kinh tế"
    """
    if not category_str:
        return None
    
    # Split by /// to get segments
    segments = category_str.split('///')
    # Return the last non-empty segment, stripped
    for seg in reversed(segments):
        seg = seg.strip()
        if seg:
            return seg
    return None


def _find_or_create_author(db: Session, author_name: str) -> Optional[Author]:
    """Find an existing author by name or create a new one."""
    if not author_name:
        return None
    
    author = db.query(Author).filter(Author.name == author_name).first()
    if not author:
        author = Author(name=author_name)
        db.add(author)
        db.flush()
        logger.info(f"Created new author: {author_name}")
    return author


def _find_or_create_category(db: Session, category_name: str) -> Optional[Category]:
    """Find an existing category by name or create a new one."""
    if not category_name:
        return None
    
    category = db.query(Category).filter(Category.name == category_name).first()
    if not category:
        category = Category(name=category_name)
        db.add(category)
        db.flush()
        logger.info(f"Created new category: {category_name}")
    return category


@router.post("/import-xlsx")
async def import_books_from_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    redis: Redis = Depends(get_redis),
    current_user: User = Depends(require_admin)
):
    """
    Import books from an XLSX file (Admin only).
    
    Expected columns: ISBN, Product name, Category, Price, Weight, Features, List price, Mô tả
    Missing fields are allowed - books will be created with defaults.
    Duplicate ISBNs are skipped.
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be an Excel file (.xlsx or .xls)"
        )
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl is not installed on the server"
        )
    
    # Read the file
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    
    # Read headers from first row
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val).strip() if val else '')
    
    # Map expected column names to indices (case-insensitive)
    col_map = {}
    expected_cols = {
        'isbn': 'isbn',
        'product name': 'product_name',
        'category': 'category',
        'price': 'price',
        'weight': 'weight',
        'features': 'features',
        'list price': 'list_price',
        'mô tả': 'description',
    }
    
    for idx, header in enumerate(headers):
        header_lower = header.lower().strip()
        if header_lower in expected_cols:
            col_map[expected_cols[header_lower]] = idx
    
    if 'product_name' not in col_map:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required column 'Product name'. Found columns: {headers}"
        )
    
    # Process rows
    created = 0
    skipped = 0
    errors = []
    
    # Get all existing ISBNs for fast duplicate check
    existing_isbns = set()
    all_isbns = db.query(Book.isbn).filter(Book.isbn.isnot(None), Book.isbn != '').all()
    for (isbn_val,) in all_isbns:
        existing_isbns.add(str(isbn_val).strip())
    
    # Get existing slugs for uniqueness check
    existing_slugs = set()
    all_slugs = db.query(Book.slug).filter(Book.slug.isnot(None)).all()
    for (slug_val,) in all_slugs:
        existing_slugs.add(slug_val)
    
    for row_num in range(2, ws.max_row + 1):
        # Use a savepoint so one row's error doesn't kill the whole transaction
        savepoint = db.begin_nested()
        try:
            # Read cell values
            def get_val(col_key):
                if col_key not in col_map:
                    return None
                val = ws.cell(row=row_num, column=col_map[col_key] + 1).value
                return val
            
            title = get_val('product_name')
            if not title:
                savepoint.rollback()
                continue  # Skip empty rows
            
            title = str(title).strip()
            isbn_raw = str(get_val('isbn') or '').strip()
            
            # Truncate ISBN to 13 chars (DB column limit) — keep only digits if too long
            isbn = isbn_raw
            if isbn and len(isbn) > 13:
                # Try extracting just the digits
                digits_only = re.sub(r'[^0-9]', '', isbn)
                isbn = digits_only[:13] if digits_only else isbn[:13]
            
            # Skip duplicate ISBN
            if isbn and isbn in existing_isbns:
                skipped += 1
                savepoint.rollback()
                continue
            
            # Parse price
            raw_price = get_val('price')
            try:
                price = int(float(str(raw_price).replace(',', '').strip())) if raw_price else 0
            except (ValueError, TypeError):
                price = 0
            
            # Parse weight
            raw_weight = get_val('weight')
            try:
                weight = int(float(str(raw_weight).replace(',', '').strip())) if raw_weight else None
            except (ValueError, TypeError):
                weight = None
            
            # Description
            description = str(get_val('description') or '').strip() or None
            
            # Parse features
            features_str = str(get_val('features') or '')
            parsed_features = _parse_features(features_str)
            
            # Parse category
            category_str = str(get_val('category') or '')
            category_name = _parse_category(category_str)
            
            # Generate slug
            slug = _slugify(title)
            if slug in existing_slugs:
                # Make unique by appending ISBN or row number
                slug = f"{slug}-{isbn}" if isbn else f"{slug}-{row_num}"
                slug = _slugify(slug)
            
            if slug in existing_slugs:
                # Still duplicate, skip
                skipped += 1
                errors.append(f"Row {row_num}: Duplicate slug for '{title[:50]}'")
                savepoint.rollback()
                continue
            
            # Create the book
            db_book = Book(
                title=title,
                slug=slug,
                isbn=isbn if isbn else None,
                price=price,
                stock_quantity=0,  # Default, admin can update later
                full_description=description,
                weight=weight,
                pages=parsed_features.get('pages'),
                height=parsed_features.get('height'),
                width=parsed_features.get('width'),
                is_active=True,
            )
            
            db.add(db_book)
            db.flush()
            
            # Find or create author
            author_name = parsed_features.get('author')
            if author_name:
                author = _find_or_create_author(db, author_name)
                if author:
                    db_book.authors.append(author)
            
            # Find or create category
            if category_name:
                category = _find_or_create_category(db, category_name)
                if category:
                    db_book.categories.append(category)
            
            # Commit the savepoint
            savepoint.commit()
            
            # Track for duplicate prevention
            if isbn:
                existing_isbns.add(isbn)
            existing_slugs.add(slug)
            
            created += 1
            
            # Batch commit every 100 rows for performance
            if created % 100 == 0:
                db.commit()
                logger.info(f"Import progress: {created} books created so far...")
        
        except Exception as e:
            savepoint.rollback()
            errors.append(f"Row {row_num}: {str(e)[:100]}")
            continue
    
    # Final commit
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database commit failed: {str(e)}"
        )
    
    # Invalidate book cache
    cache = RedisCache(redis)
    await cache.delete_pattern("books:*")
    
    wb.close()
    
    logger.info(f"Import completed: {created} created, {skipped} skipped, {len(errors)} errors")
    
    return {
        "message": f"Import completed successfully",
        "created": created,
        "skipped": skipped,
        "total_errors": len(errors),
        "errors": errors[:50]  # Return first 50 errors to avoid huge response
    }
